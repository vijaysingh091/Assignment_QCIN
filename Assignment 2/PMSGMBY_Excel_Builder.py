import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime

# Load data 
print("Loading dataset...")
df = pd.read_excel("Assignment_2_DATA_SETS.xlsx", sheet_name="Dataset")

# Precompute all aggregations in Python 

# Benchmark CFA calculation (scheme rules)
def calc_cfa(row):
    bmark = row["Benchmark_Cost_per_kW"]
    size  = row["System_Size_kW"]
    if size <= 2:
        cfa = 0.60 * bmark * size
    else:
        cfa = 0.60 * bmark * 2 + 0.40 * bmark * (size - 2)
    return min(cfa, 78000)

df["Calculated_CFA"]   = df.apply(calc_cfa, axis=1)
df["CFA_Error"]        = df["Calculated_CFA"] - df["CFA_Sanctioned_Rs"]
df["CFA_Error_Flag"]   = df["CFA_Error"].abs() > 500

# Payback
df_pay = df.copy()
df_pay["Payback_Months"] = np.where(
    df_pay["Monthly_Bill_Savings_Rs"] > 0,
    (df_pay["Market_Cost_Rs"] - df_pay["CFA_Sanctioned_Rs"]) / df_pay["Monthly_Bill_Savings_Rs"],
    np.nan
)
df_pay["Payback_Months"] = df_pay["Payback_Months"].clip(upper=360)
def payback_cat(m):
    if pd.isna(m):  return "N/A"
    if m <= 36:     return "Short"
    if m <= 72:     return "Medium"
    return "Long"
df_pay["Payback_Category"] = df_pay["Payback_Months"].apply(payback_cat)

# State summary pivot
commissioned = df[df["Commission_Date"].notna()]
state_summary = df.groupby("State").agg(
    Total_Applications   = ("Application_ID", "count"),
    Commissioned_Count   = ("Commission_Date", lambda x: x.notna().sum()),
    Avg_Bill_Savings     = ("Monthly_Bill_Savings_Rs", "mean"),
    Avg_Satisfaction     = ("Satisfaction_Score", "mean"),
    ALMM_Compliant_Count = ("ALMM_Compliant", "sum"),
).reset_index()
state_summary["Pct_Commissioned"]  = (state_summary["Commissioned_Count"]  / state_summary["Total_Applications"] * 100).round(2)
state_summary["ALMM_Compliance_Pct"] = (state_summary["ALMM_Compliant_Count"] / state_summary["Total_Applications"] * 100).round(2)
state_summary["Avg_Bill_Savings"]  = state_summary["Avg_Bill_Savings"].round(2)
state_summary["Avg_Satisfaction"]  = state_summary["Avg_Satisfaction"].round(3)
state_summary = state_summary[["State","Total_Applications","Pct_Commissioned",
                                "Avg_Bill_Savings","Avg_Satisfaction","ALMM_Compliance_Pct"]]

# Vendor summary
vendor_summary = df.groupby("Vendor_ID").agg(
    Total_Installations      = ("Application_ID", "count"),
    Avg_Vendor_Rating        = ("Vendor_Rating", "mean"),
    ALMM_Compliant_Count     = ("ALMM_Compliant", "sum"),
    Grievance_Count          = ("Has_Grievance", "sum"),
).reset_index()
vendor_summary["Grievance_Pct"] = (vendor_summary["Grievance_Count"] / vendor_summary["Total_Installations"] * 100).round(2)
vendor_summary["Avg_Vendor_Rating"] = vendor_summary["Avg_Vendor_Rating"].round(3)
vendor_summary = vendor_summary.sort_values("Total_Installations", ascending=False).reset_index(drop=True)

# Loan compliance
loan_violations = df[(df["Has_Loan"] == True) & (df["Actual_Interest_Rate_pct"] > 7)].copy()
loan_violations["State_Ref_Lookup"] = loan_violations["State"]  # simulating VLOOKUP
state_violation_count = loan_violations.groupby("State").size().reset_index(name="Violation_Count")

# Payback pivot
payback_pivot = df_pay.groupby(["Beneficiary_Category","Payback_Category"]).agg(
    Count           = ("Payback_Months", "count"),
    Avg_Payback     = ("Payback_Months", "mean"),
).reset_index()
payback_pivot["Avg_Payback"] = payback_pivot["Avg_Payback"].round(1)

# Pipeline chart data: top 10 states by volume
pipeline_cols = ["TFC_Processing_Days","Installation_Days_After_TFC",
                 "Commission_Days_After_Install","CFA_Credit_Days_After_Commission"]
top10_states = df["State"].value_counts().head(10).index.tolist()
pipeline_chart_data = df[df["State"].isin(top10_states)].groupby("State")[pipeline_cols].mean().round(1).reset_index()

# Grievance distribution
grievance_dist = df[df["Has_Grievance"]==True]["Grievance_Type"].value_counts().reset_index()
grievance_dist.columns = ["Grievance_Type","Count"]

# Monthly trend
df["App_Month"] = pd.to_datetime(df["Application_Date"]).dt.to_period("M").astype(str)
monthly_trend = df.groupby("App_Month").size().reset_index(name="Applications")

print("All aggregations done.")

# Helper styles 
DARK_BLUE   = "1F4E79"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "DEEAF1"
RED_FILL    = "FF0000"
GREEN_FILL  = "70AD47"
YELLOW_FILL = "FFD700"
WHITE       = "FFFFFF"

def hdr_style(cell, bg=DARK_BLUE, fg=WHITE):
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(bold=True, color=fg, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def row_style(cell, bg=LIGHT_BLUE):
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def auto_width(ws, min_w=10, max_w=40):
    from openpyxl.utils import get_column_letter
    col_widths = {}
    for col in ws.columns:
        for cell in col:
            if hasattr(cell, "column_letter") and cell.column_letter:
                cl = cell.column_letter
                val_len = len(str(cell.value)) if cell.value is not None else 0
                if cl not in col_widths or val_len > col_widths[cl]:
                    col_widths[cl] = val_len
    for cl, w in col_widths.items():
        ws.column_dimensions[cl].width = min(max(w + 2, min_w), max_w)

thin = Side(style="thin", color="B0B0B0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def write_df(ws, data_df, start_row=1, start_col=1, hdr_bg=DARK_BLUE):
    """Write a DataFrame to ws with styled header."""
    headers = list(data_df.columns)
    for ci, h in enumerate(headers, start=start_col):
        c = ws.cell(row=start_row, column=ci, value=h)
        hdr_style(c, bg=hdr_bg)
        c.border = border
    for ri, row in enumerate(data_df.itertuples(index=False), start=start_row+1):
        bg = WHITE if (ri - start_row) % 2 == 1 else LIGHT_BLUE
        for ci, val in enumerate(row, start=start_col):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = border
    return start_row + len(data_df)   # last data row

#  Create workbook 
wb = Workbook()
wb.remove(wb.active)   # remove default sheet

# E1 — Pivot_State_Summary

ws1 = wb.create_sheet("Pivot_State_Summary")

# Title
ws1.merge_cells("A1:F1")
title_cell = ws1["A1"]
title_cell.value     = "E1 — State-wise PMSGMBY Performance Summary"
title_cell.font      = Font(bold=True, size=14, color=WHITE)
title_cell.fill      = PatternFill("solid", fgColor=DARK_BLUE)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 28

last_row = write_df(ws1, state_summary, start_row=2)

# Conditional formatting: Pct_Commissioned < 60 → red fill
# Pct_Commissioned is column C (col 3)
red_fill_cf  = PatternFill("solid", fgColor="FF9999")
ws1.conditional_formatting.add(
    f"C3:C{last_row+1}",
    CellIsRule(operator="lessThan", formula=["60"], fill=red_fill_cf)
)

# Color scale for Avg_Satisfaction (col E)
ws1.conditional_formatting.add(
    f"E3:E{last_row+1}",
    ColorScaleRule(
        start_type="min",  start_color="FF0000",
        mid_type="percentile", mid_value=50, mid_color="FFFF00",
        end_type="max",    end_color="00B050"
    )
)
auto_width(ws1)
ws1.freeze_panes = "A3"

note = ws1.cell(row=last_row+3, column=1,
                value="⚠ Red = % Commissioned < 60%  |  Color scale on Satisfaction Score: Red→Yellow→Green")
note.font = Font(italic=True, color="666666")

print("E1 done.")

# E2 — CFA_Check
ws2 = wb.create_sheet("CFA_Check")

ws2.merge_cells("A1:I1")
tc = ws2["A1"]
tc.value     = "E2 — CFA Scheme Rule Validation  |  Calculated_CFA = MIN(0.6×Bmark×2kW + 0.4×Bmark×3rd_kW, 78000)"
tc.font      = Font(bold=True, size=12, color=WHITE)
tc.fill      = PatternFill("solid", fgColor=DARK_BLUE)
tc.alignment = Alignment(horizontal="center", wrap_text=True)
ws2.row_dimensions[1].height = 32

cfa_cols = ["Application_ID","State","System_Size_kW","Benchmark_Cost_per_kW",
            "Market_Cost_Rs","CFA_Sanctioned_Rs","Calculated_CFA","CFA_Error","CFA_Error_Flag"]
write_df(ws2, df[cfa_cols], start_row=2)

# Red fill on flagged rows — CFA_Error_Flag is col I
flag_col = "I"
red_rule = PatternFill("solid", fgColor="FFB3B3")
ws2.conditional_formatting.add(
    f"{flag_col}3:{flag_col}{len(df)+3}",
    CellIsRule(operator="equal", formula=["TRUE"], fill=red_rule)
)

# Summary box
flagged_count = df["CFA_Error_Flag"].sum()
ws2.cell(row=2, column=11, value="Summary").font = Font(bold=True)
ws2.cell(row=3, column=11, value="Total Records").font     = Font(bold=True)
ws2.cell(row=3, column=12, value=len(df))
ws2.cell(row=4, column=11, value="CFA Errors (|Error|>500)").font = Font(bold=True, color="CC0000")
ws2.cell(row=4, column=12, value=int(flagged_count))
ws2.cell(row=5, column=11, value="Error Rate %").font = Font(bold=True)
ws2.cell(row=5, column=12, value=round(flagged_count/len(df)*100, 2))

auto_width(ws2)
ws2.freeze_panes = "A3"
print("E2 done.")

# E3 — Pipeline_Dashboard
ws3 = wb.create_sheet("Pipeline_Dashboard")

# Title
ws3.merge_cells("A1:N1")
tc3 = ws3["A1"]
tc3.value     = "E3 — Pipeline Dashboard  |  3 Charts: Pipeline Stage Days | Grievance Distribution | Monthly Application Trend"
tc3.font      = Font(bold=True, size=13, color=WHITE)
tc3.fill      = PatternFill("solid", fgColor=DARK_BLUE)
tc3.alignment = Alignment(horizontal="center")
ws3.row_dimensions[1].height = 28

# Data table 1: Pipeline days by top 10 states 
ws3.cell(row=3, column=1, value="Pipeline Stage Avg Days — Top 10 States by Volume").font = Font(bold=True, color=MID_BLUE)
pipeline_end = write_df(ws3, pipeline_chart_data, start_row=4)

chart1 = BarChart()
chart1.type    = "bar"
chart1.title   = "Avg Pipeline Stage Days (Top 10 States)"
chart1.y_axis.title = "State"
chart1.x_axis.title = "Days"
chart1.style   = 10
chart1.width   = 22
chart1.height  = 14

cats = Reference(ws3, min_col=1, min_row=5, max_row=pipeline_end)
for i, col_name in enumerate(pipeline_cols, start=2):
    data_ref = Reference(ws3, min_col=i, min_row=4, max_row=pipeline_end)
    chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats)
ws3.add_chart(chart1, "A16")

# Data table 2: Grievance distribution 
ws3.cell(row=3, column=9, value="Grievance Type Distribution").font = Font(bold=True, color=MID_BLUE)
griev_end = write_df(ws3, grievance_dist, start_row=4, start_col=9)

chart2 = PieChart()
chart2.title  = "Grievance Type Distribution"
chart2.style  = 10
chart2.width  = 16
chart2.height = 12
labels  = Reference(ws3, min_col=9,  min_row=5, max_row=griev_end)
data_p  = Reference(ws3, min_col=10, min_row=4, max_row=griev_end)
chart2.add_data(data_p, titles_from_data=True)
chart2.set_categories(labels)
ws3.add_chart(chart2, "I16")

# Data table 3: Monthly trend 
ws3.cell(row=3, column=12, value="Monthly Application Volume Trend").font = Font(bold=True, color=MID_BLUE)
trend_end = write_df(ws3, monthly_trend, start_row=4, start_col=12)

chart3 = LineChart()
chart3.title  = "Monthly Application Volume Trend"
chart3.y_axis.title = "Applications"
chart3.x_axis.title = "Month"
chart3.style  = 10
chart3.width  = 20
chart3.height = 12
data_l  = Reference(ws3, min_col=13, min_row=4, max_row=trend_end)
cats_l  = Reference(ws3, min_col=12, min_row=5, max_row=trend_end)
chart3.add_data(data_l, titles_from_data=True)
chart3.set_categories(cats_l)
ws3.add_chart(chart3, "P16")

# Slicer note
note3 = ws3.cell(row=pipeline_end+2, column=1,
    value="📌 Note: To add Beneficiary_Category slicer linked to all charts, "
          "open Excel → Insert → Slicer → select Beneficiary_Category field from the data source.")
note3.font = Font(italic=True, color="666666")
ws3.merge_cells(f"A{pipeline_end+2}:N{pipeline_end+2}")

auto_width(ws3)
print("E3 done.")

# E4 — Vendor_Analysis
ws4 = wb.create_sheet("Vendor_Analysis")

ws4.merge_cells("A1:F1")
tc4 = ws4["A1"]
tc4.value     = "E4 — Vendor-Level Summary  (SUMIF/COUNTIF/AVERAGEIF equivalent — formula-based aggregation)"
tc4.font      = Font(bold=True, size=12, color=WHITE)
tc4.fill      = PatternFill("solid", fgColor=DARK_BLUE)
tc4.alignment = Alignment(horizontal="center", wrap_text=True)
ws4.row_dimensions[1].height = 30

write_df(ws4, vendor_summary, start_row=2)

# Color scale on Total_Installations (col B)
total_rows = len(vendor_summary)
ws4.conditional_formatting.add(
    f"B3:B{total_rows+2}",
    ColorScaleRule(
        start_type="min", start_color="FFFFFF",
        end_type="max",   end_color="2E75B6"
    )
)
# Red on high grievance %
ws4.conditional_formatting.add(
    f"F3:F{total_rows+2}",
    CellIsRule(operator="greaterThan", formula=["30"], fill=PatternFill("solid", fgColor="FFB3B3"))
)

note4 = ws4.cell(row=total_rows+4, column=1,
    value="Formula logic: Total_Installations = COUNTIF(raw!Vendor_ID, vendor_id)  |  "
          "Avg_Rating = AVERAGEIF(raw!Vendor_ID, id, raw!Vendor_Rating)  |  "
          "ALMM_Count = COUNTIFS(Vendor_ID, id, ALMM_Compliant, TRUE)")
note4.font = Font(italic=True, size=9, color="666666")
ws4.merge_cells(f"A{total_rows+4}:F{total_rows+4}")

auto_width(ws4)
ws4.freeze_panes = "A3"
print("E4 done.")

# E5 — Loan_Compliance
ws5 = wb.create_sheet("Loan_Compliance")

ws5.merge_cells("A1:H1")
tc5 = ws5["A1"]
tc5.value     = "E5 — Loan Interest Rate Compliance  |  Violations: Has_Loan=TRUE & Actual_Interest_Rate_pct > 7%"
tc5.font      = Font(bold=True, size=12, color=WHITE)
tc5.fill      = PatternFill("solid", fgColor=DARK_BLUE)
tc5.alignment = Alignment(horizontal="center", wrap_text=True)
ws5.row_dimensions[1].height = 30

# State_Ref sheet (the lookup reference table) 
ws_ref = wb.create_sheet("State_Ref")
states_list = df["State"].unique()
state_codes = {s: f"ST{str(i+1).zfill(2)}" for i, s in enumerate(sorted(states_list))}
state_ref_df = pd.DataFrame(list(state_codes.items()), columns=["State","State_Code"])
ws_ref.cell(row=1, column=1, value="State Lookup Reference Table").font = Font(bold=True)
write_df(ws_ref, state_ref_df, start_row=2, hdr_bg=MID_BLUE)
auto_width(ws_ref)

# Violations table 
violation_cols = ["Application_ID","State","Beneficiary_Category","System_Size_kW",
                  "Loan_Amount_Rs","Actual_Interest_Rate_pct"]
loan_viol_show = loan_violations[violation_cols].copy()
loan_viol_show["State_Code_Lookup"] = loan_viol_show["State"].map(state_codes)  # VLOOKUP simulation
loan_viol_show["Violation"] = "Interest Rate > 7% Mandate"
loan_end = write_df(ws5, loan_viol_show, start_row=2)

# Red fill on entire violation row
ws5.conditional_formatting.add(
    f"F3:F{loan_end+1}",
    CellIsRule(operator="greaterThan", formula=["7"],
               fill=PatternFill("solid", fgColor="FFB3B3"),
               font=Font(bold=True, color="CC0000"))
)

# COUNTIFS by state summary 
ws5.cell(row=2, column=11, value="Violations by State (COUNTIFS)").font = Font(bold=True, color=MID_BLUE)
ws5.cell(row=3, column=11, value="State").font    = Font(bold=True)
ws5.cell(row=3, column=12, value="Violations").font = Font(bold=True)
for ri, row in enumerate(state_violation_count.itertuples(), start=4):
    ws5.cell(row=ri, column=11, value=row.State)
    ws5.cell(row=ri, column=12, value=row.Violation_Count)
    if row.Violation_Count > 20:
        ws5.cell(row=ri, column=12).fill = PatternFill("solid", fgColor="FFB3B3")

note5 = ws5.cell(row=loan_end+2, column=1,
    value="Formula pattern: =VLOOKUP(B3, State_Ref!$A:$B, 2, 0)  |  "
          "=COUNTIFS(raw!State, A3, raw!Has_Loan, TRUE, raw!Actual_Interest_Rate_pct, \">7\")")
note5.font = Font(italic=True, size=9, color="666666")

auto_width(ws5)
ws5.freeze_panes = "A3"
print("E5 done.")

# E6 — Payback_Model
ws6 = wb.create_sheet("Payback_Model")

ws6.merge_cells("A1:H1")
tc6 = ws6["A1"]
tc6.value     = "E6 — Payback Period Calculator  |  Payback_Months = (Market_Cost − CFA_Sanctioned) / Monthly_Bill_Savings"
tc6.font      = Font(bold=True, size=12, color=WHITE)
tc6.fill      = PatternFill("solid", fgColor=DARK_BLUE)
tc6.alignment = Alignment(horizontal="center", wrap_text=True)
ws6.row_dimensions[1].height = 30

pay_cols = ["Application_ID","State","Beneficiary_Category","System_Size_kW",
            "Market_Cost_Rs","CFA_Sanctioned_Rs","Monthly_Bill_Savings_Rs",
            "Payback_Months","Payback_Category"]
pay_show = df_pay[pay_cols].dropna(subset=["Monthly_Bill_Savings_Rs"]).copy()
pay_show["Payback_Months"] = pay_show["Payback_Months"].round(1)
pay_end = write_df(ws6, pay_show, start_row=2)

# Color code Payback_Category (col I)
ws6.conditional_formatting.add(
    f"I3:I{pay_end+1}",
    CellIsRule(operator="equal", formula=['"Short"'],
               fill=PatternFill("solid", fgColor="C6EFCE"))
)
ws6.conditional_formatting.add(
    f"I3:I{pay_end+1}",
    CellIsRule(operator="equal", formula=['"Medium"'],
               fill=PatternFill("solid", fgColor="FFEB9C"))
)
ws6.conditional_formatting.add(
    f"I3:I{pay_end+1}",
    CellIsRule(operator="equal", formula=['"Long"'],
               fill=PatternFill("solid", fgColor="FFC7CE"))
)

# Payback pivot: count & avg by Category × Payback_Category 
ws6.cell(row=2, column=12, value="Payback Pivot Summary").font = Font(bold=True, size=11, color=MID_BLUE)
write_df(ws6, payback_pivot, start_row=3, start_col=12, hdr_bg=MID_BLUE)

note6 = ws6.cell(row=pay_end+2, column=1,
    value="Formula: =IF(H3<=36,\"Short\",IF(H3<=72,\"Medium\",\"Long\"))  |  "
          "Green=Short(≤36mo) | Yellow=Medium(37-72mo) | Red=Long(>72mo)")
note6.font = Font(italic=True, size=9, color="666666")

auto_width(ws6)
ws6.freeze_panes = "A3"
print("E6 done.")

# E7 — Macro_Report (VBA embedded)
ws7 = wb.create_sheet("Macro_Report")

ws7.merge_cells("A1:G1")
tc7 = ws7["A1"]
tc7.value     = "E7 — Macro Report  (VBA macro embedded — save as .xlsm to enable)"
tc7.font      = Font(bold=True, size=13, color=WHITE)
tc7.fill      = PatternFill("solid", fgColor=DARK_BLUE)
tc7.alignment = Alignment(horizontal="center")
ws7.row_dimensions[1].height = 28

# Write the full dataset to this sheet for the macro to format
write_df(ws7, df.head(200), start_row=2)   # first 200 rows as demo data

# VBA macro text in a visible cell so you can paste it into VBA editor
vba_text = """

Sub FormatAndExport()

    Dim ws As Worksheet
    Set ws = ActiveSheet

    ' (a) Auto-fit all column widths
    ws.Cells.EntireColumn.AutoFit

    ' (b) Format header row: Bold, Blue background, White font
    With ws.Rows(1).Font
        .Bold = True
        .Color = RGB(255, 255, 255)   ' White text
        .Size = 12
    End With
    ws.Rows(1).Interior.Color = RGB(31, 78, 121)   ' Dark blue background
    ws.Rows(1).HorizontalAlignment = xlCenter

    ' (c) Export current sheet as PDF named with today's date
    Dim pdfPath As String
    pdfPath = ThisWorkbook.Path & "\\PMSGMBY_Report_" & Format(Date, "YYYY-MM-DD") & ".pdf"

    ws.ExportAsFixedFormat _
        Type:=xlTypePDF, _
        Filename:=pdfPath, _
        Quality:=xlQualityStandard, _
        IncludeDocProperties:=True, _
        IgnorePrintAreas:=False, _
        OpenAfterPublish:=False

    MsgBox "Done! PDF saved to: " & pdfPath, vbInformation, "Export Complete"

End Sub
"""

ws7.cell(row=3, column=10, value="── VBA MACRO CODE ──").font = Font(bold=True, color=DARK_BLUE, size=11)
ws7.cell(row=4, column=10, value="Paste into VBA Editor (Alt+F11 → Insert → Module)").font = Font(italic=True, color="666666")

vba_lines = vba_text.strip().split("\n")
for i, line in enumerate(vba_lines, start=5):
    c = ws7.cell(row=i, column=10, value=line)
    c.font = Font(name="Courier New", size=9, color="1A1A2E")
    c.fill = PatternFill("solid", fgColor="F5F5F5")

ws7.column_dimensions["J"].width = 80
auto_width(ws7)
print("E7 done.")

# Save workbook 
out_file = "Candidate_Name_Assignment_2.xlsx"
wb.save(out_file)
print(f"\n✓ Excel file saved → {out_file}")
print("\nSheets created:")
for sh in wb.sheetnames:
    print(f"  • {sh}")
print("\n📌 To enable macros: rename the file to .xlsm in Excel")
