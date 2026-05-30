import sqlite3
import pandas as pd
import numpy as np

# 1. Load dataset 
print("Loading dataset...")
df = pd.read_excel("Assignment_2_DATA_SETS.xlsx", sheet_name="Dataset")

# SQLite needs booleans as integers
for col in df.select_dtypes(include="bool").columns:
    df[col] = df[col].astype(int)

# SQLite needs dates as strings
for col in df.select_dtypes(include=["datetime64"]).columns:
    df[col] = df[col].dt.strftime("%Y-%m-%d")

print(f"  {len(df):,} rows loaded.")

# 2. Create in-memory SQLite DB 
con = sqlite3.connect(":memory:")
df.to_sql("pmsgmby_applications", con, index=False, if_exists="replace")
print("  Table 'pmsgmby_applications' created in SQLite.")

# ── 3. Define all 8 queries 
queries = {}

queries["S1_State_Conversion"] = """
SELECT
    State,
    COUNT(*)                                                          AS Total_Applications,
    SUM(CASE WHEN Commission_Date IS NOT NULL THEN 1 ELSE 0 END)     AS Commissioned,
    ROUND(
        100.0 * SUM(CASE WHEN Commission_Date IS NOT NULL THEN 1 ELSE 0 END)
              / COUNT(*), 2
    )                                                                 AS Conversion_Rate_Pct
FROM pmsgmby_applications
GROUP BY State
HAVING COUNT(*) >= 50
ORDER BY Conversion_Rate_Pct DESC;
"""

queries["S2_DISCOM_DBT_Breach"] = """
SELECT
    DISCOM,
    State,
    ROUND(AVG(CFA_Credit_Days_After_Commission), 2) AS Avg_CFA_Days,
    COUNT(*)                                         AS Applications_In_Breach
FROM pmsgmby_applications
WHERE CFA_Credit_Days_After_Commission IS NOT NULL
GROUP BY DISCOM, State
HAVING AVG(CFA_Credit_Days_After_Commission) > 60
ORDER BY Avg_CFA_Days DESC;
"""

queries["S3_Top_Vendors_Per_State"] = """
WITH vendor_stats AS (
    SELECT
        State,
        Vendor_ID,
        ROUND(AVG(Vendor_Rating), 3)  AS Avg_Rating,
        COUNT(*)                       AS Num_Installations,
        DENSE_RANK() OVER (
            PARTITION BY State
            ORDER BY AVG(Vendor_Rating) DESC
        )                              AS Rating_Rank
    FROM pmsgmby_applications
    WHERE Vendor_Rating IS NOT NULL
      AND Commission_Date IS NOT NULL
    GROUP BY State, Vendor_ID
)
SELECT State, Vendor_ID, Avg_Rating, Num_Installations, Rating_Rank
FROM vendor_stats
WHERE Rating_Rank <= 3
ORDER BY State, Rating_Rank;
"""

queries["S4_Monthly_Volume_Growth"] = """
WITH monthly AS (
    SELECT
        strftime('%Y-%m', Application_Date)  AS Year_Month,
        COUNT(*)                              AS Monthly_Volume
    FROM pmsgmby_applications
    GROUP BY Year_Month
),
cumulative AS (
    SELECT
        Year_Month,
        Monthly_Volume,
        SUM(Monthly_Volume) OVER (ORDER BY Year_Month)  AS Cumulative_Total,
        LAG(Monthly_Volume)  OVER (ORDER BY Year_Month)  AS Prev_Month_Volume
    FROM monthly
)
SELECT
    Year_Month,
    Monthly_Volume,
    Cumulative_Total,
    Prev_Month_Volume,
    CASE
        WHEN Prev_Month_Volume IS NOT NULL AND Prev_Month_Volume > 0
        THEN ROUND(100.0 * (Monthly_Volume - Prev_Month_Volume) / Prev_Month_Volume, 2)
        ELSE NULL
    END AS MoM_Growth_Rate_Pct
FROM cumulative
ORDER BY Year_Month;
"""

queries["S5_Grievance_Summary"] = """
SELECT
    Grievance_Type,
    COUNT(*)                                                          AS Total_Count,
    SUM(CASE WHEN Grievance_Resolved = 1 THEN 1 ELSE 0 END)         AS Resolved_Count,
    ROUND(100.0 * SUM(CASE WHEN Grievance_Resolved = 1 THEN 1 ELSE 0 END)
                / COUNT(*), 2)                                        AS Resolution_Rate_Pct,
    ROUND(AVG(Grievance_TAT_Days), 1)                                 AS Avg_TAT_Days,
    CASE
        WHEN ROUND(100.0 * SUM(CASE WHEN Grievance_Resolved = 1 THEN 1 ELSE 0 END)
                         / COUNT(*), 2) < 60
          OR ROUND(AVG(Grievance_TAT_Days), 1) > 30
        THEN 'Requires Attention'
        ELSE 'On Track'
    END                                                               AS SLA_Flag
FROM pmsgmby_applications
WHERE Has_Grievance = 1
  AND Grievance_Type IS NOT NULL
GROUP BY Grievance_Type
ORDER BY Resolution_Rate_Pct ASC;
"""

queries["S6_DISCOM_Incentive_Gap"] = """
WITH incentive_gap AS (
    SELECT
        DISCOM,
        SUM(DISCOM_Incentive_Received_Rs)              AS Total_Received,
        SUM(DISCOM_Incentive_Utilised_Rs)              AS Total_Utilised,
        SUM(DISCOM_Incentive_Received_Rs
            - DISCOM_Incentive_Utilised_Rs)            AS Unutilised_Gap
    FROM pmsgmby_applications
    WHERE DISCOM_Incentive_Received_Rs IS NOT NULL
      AND DISCOM_Incentive_Utilised_Rs IS NOT NULL
    GROUP BY DISCOM
),
tfc_perf AS (
    SELECT
        DISCOM,
        ROUND(AVG(TFC_Processing_Days), 2) AS Avg_TFC_Days
    FROM pmsgmby_applications
    WHERE TFC_Processing_Days IS NOT NULL
    GROUP BY DISCOM
)
SELECT
    i.DISCOM,
    ROUND(i.Total_Received, 2)   AS Total_Incentive_Received,
    ROUND(i.Total_Utilised, 2)   AS Total_Incentive_Utilised,
    ROUND(i.Unutilised_Gap, 2)   AS Unutilised_Gap,
    t.Avg_TFC_Days
FROM incentive_gap i
LEFT JOIN tfc_perf t ON i.DISCOM = t.DISCOM
ORDER BY i.Unutilised_Gap DESC
LIMIT 10;
"""

queries["S7_Loan_Interest_Brackets"] = """
WITH bracketed AS (
    SELECT
        Application_ID,
        State,
        Loan_Amount_Rs,
        System_Size_kW,
        Actual_Interest_Rate_pct,
        CASE
            WHEN Actual_Interest_Rate_pct <= 7 THEN 'Compliant (<=7%)'
            WHEN Actual_Interest_Rate_pct <= 9 THEN 'Marginal (7-9%)'
            ELSE 'Non-Compliant (>9%)'
        END AS Rate_Bracket
    FROM pmsgmby_applications
    WHERE Has_Loan = 1
      AND Actual_Interest_Rate_pct IS NOT NULL
),
bracket_agg AS (
    SELECT
        Rate_Bracket,
        COUNT(*)                           AS Count,
        ROUND(AVG(Loan_Amount_Rs), 0)      AS Avg_Loan_Amount,
        ROUND(AVG(System_Size_kW), 2)      AS Avg_System_Size_kW
    FROM bracketed
    GROUP BY Rate_Bracket
),
top_state AS (
    SELECT
        Rate_Bracket,
        State,
        ROW_NUMBER() OVER (
            PARTITION BY Rate_Bracket
            ORDER BY COUNT(*) DESC
        ) AS rn
    FROM bracketed
    GROUP BY Rate_Bracket, State
)
SELECT
    b.Rate_Bracket,
    b.Count,
    b.Avg_Loan_Amount,
    b.Avg_System_Size_kW,
    t.State AS State_With_Highest_Count
FROM bracket_agg b
LEFT JOIN top_state t ON b.Rate_Bracket = t.Rate_Bracket AND t.rn = 1
ORDER BY b.Count DESC;
"""

queries["S8_High_Value_Underserved"] = """
SELECT
    State,
    COUNT(*)                                              AS Underserved_Count,
    ROUND(AVG(Market_Cost_Rs - CFA_Sanctioned_Rs), 2)    AS Avg_Out_Of_Pocket
FROM pmsgmby_applications
WHERE Beneficiary_Category IN ('BPL', 'SC/ST')
  AND System_Size_kW = 3
  AND (Loan_Approved IS NULL OR Loan_Approved = 0)
  AND CFA_Sanctioned_Rs < 0.5 * Market_Cost_Rs
GROUP BY State
ORDER BY Underserved_Count DESC;
"""

# 4. Run all queries and collect results 
results = {}
print("\n" + "=" * 60)
for qname, sql in queries.items():
    result_df = pd.read_sql_query(sql, con)
    results[qname] = result_df
    print(f"\n{'='*60}\n{qname}\n{'='*60}")
    print(result_df.to_string(index=False))

# 5. Export all results to Excel 
out_file = "PMSGMBY_SQL_Results.xlsx"
with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
    for qname, result_df in results.items():
        sheet_name = qname[:31]   # Excel sheet name max 31 chars
        result_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Basic header formatting
        ws = writer.sheets[sheet_name]
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

con.close()
print(f"\n✓ All SQL results saved → {out_file}")
print("\nSheets written:")
for qname in queries:
    print(f"  • {qname[:31]}")
